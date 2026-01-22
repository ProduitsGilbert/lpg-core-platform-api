import io
import os

import pytest
from openpyxl import load_workbook

pytestmark = pytest.mark.integration

DEFAULT_SALES_ORDERS = [
    "GI024770",
    "GI024764",
    "GI024763",
    "GI024762",
    "GI024761",
    "GI024758",
    "GI024756",
    "GI024755",
    "GI024746",
    "GI024745",
    "GI024744",
    "GI024743",
    "GI024741",
    "GI024730",
    "GI024729",
    "GI024726",
]

SALES_ORDERS_ENV = os.environ.get("MRP_TEST_SALES_ORDERS", "")
SALES_ORDERS = [
    value.strip().upper()
    for value in SALES_ORDERS_ENV.split(",")
    if value.strip()
] or DEFAULT_SALES_ORDERS
MIN_ROWS = int(os.environ.get("MRP_TEST_MIN_ROWS", "1"))


def _require_mrp_env() -> None:
    if os.getenv("MRP_RUN_INTEGRATION") != "1":
        pytest.skip("Set MRP_RUN_INTEGRATION=1 to run live BasicMRP integration tests")
    if not SALES_ORDERS:
        pytest.skip("MRP_TEST_SALES_ORDERS is not configured")


def _mrp_url(base_url: str, path: str) -> str:
    return f"{base_url}/api/v1/mrp{path}"


def _count_child_rows(worksheet) -> int:
    count = 0
    for row in worksheet.iter_rows(min_row=5, min_col=14, max_col=23, values_only=True):
        if not row:
            continue
        if row[0] in (None, "", "Item"):
            continue
        if any(cell not in (None, "") for cell in row):
            count += 1
    return count


def test_missing_parts_export_live(http_session, base_url, default_timeout) -> None:
    _require_mrp_env()
    failures = []

    for sales_order_no in SALES_ORDERS:
        response = http_session.get(
            _mrp_url(base_url, f"/sales-orders/{sales_order_no}/missing-parts/export"),
            timeout=default_timeout,
        )

        if response.status_code != 200:
            failures.append(f"{sales_order_no}: HTTP {response.status_code}")
            continue

        content_type = response.headers.get("content-type", "")
        if "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" not in content_type:
            failures.append(f"{sales_order_no}: invalid content-type {content_type}")
            continue

        workbook = load_workbook(io.BytesIO(response.content))
        missing = [name for name in ("Export", "MRP Out", "MRP In") if name not in workbook.sheetnames]
        if missing:
            failures.append(f"{sales_order_no}: missing sheets {', '.join(missing)}")
            continue

        data_rows = _count_child_rows(workbook["Export"])
        if data_rows < MIN_ROWS:
            failures.append(
                f"{sales_order_no}: expected >= {MIN_ROWS} data rows, got {data_rows}"
            )

    assert not failures, "Missing parts export failures:\n" + "\n".join(failures)

