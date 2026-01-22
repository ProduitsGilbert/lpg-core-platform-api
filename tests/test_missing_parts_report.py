from unittest.mock import AsyncMock

import httpx
import pytest
from openpyxl import load_workbook

from app.domain.toolkit.missing_parts_service import MissingPartsReportService
from app.main import app


SAMPLE_OUT_ROWS = [
    {
        "$type": "GilbertAPI.Model.MRP.outProdOrd, GilbertAPI",
        "name": "M176567",
        "prodOrderNo": "M176567",
        "jobNo": "GI022960",
        "itemNo": "8218725",
        "qtyToFill": 2,
        "qtyFilled": 0,
        "remainingQuantity": 2,
        "description": "MEC DOUILLE FIXATION ASS.",
        "needDate": "2025-09-29T00:00:00",
        "binCode": "TO-PROD",
        "mrpTakeFrom": "TO-PROD;",
    },
    {
        "$type": "GilbertAPI.Model.MRP.outProdOrd, GilbertAPI",
        "name": "M176567",
        "prodOrderNo": "M176567",
        "jobNo": "GI022960",
        "itemNo": "622411",
        "qtyToFill": 16,
        "qtyFilled": 0,
        "remainingQuantity": 16,
        "description": "BOUL. CYL. SPC 1/4-20UNC X 1/2",
        "needDate": "2025-08-07T11:33:00",
        "binCode": "141B1",
        "mrpTakeFrom": "Stock",
        "attributionFromIn": [{"qty": 4, "name": "Stock"}],
    },
    {
        "$type": "GilbertAPI.Model.MRP.outProdOrd, GilbertAPI",
        "name": "M176567",
        "prodOrderNo": "M176567",
        "jobNo": "GI022960",
        "itemNo": "8315108",
        "qtyToFill": 2,
        "qtyFilled": 0,
        "remainingQuantity": 2,
        "description": "DOUILLE DE FIXATION",
        "needDate": "2025-08-07T11:33:00",
        "mrpTakeFrom": "1xM172792",
        "attributionFromIn": [{"qty": 1, "name": "M172792", "job": "GI022960"}],
    },
]

SAMPLE_IN_ROWS = [
    {
        "$type": "GilbertAPI.Model.MRP.InStock, GilbertAPI",
        "itemNo": "622411",
        "description": "BOUL. CYL. SPC 1/4-20UNC X 1/2",
        "qtyDisponible": 40,
        "binCode": "141B1",
        "name": "Stock",
        "dateGet": "2025-07-15T00:00:00",
    },
    {
        "$type": "GilbertAPI.Model.MRP.InProd, GilbertAPI",
        "prodOrderNo": "M176567",
        "name": "M176567",
        "itemNo": "8218725",
        "description": "MEC DOUILLE FIXATION ASS.",
        "quantity": 2,
        "remainingQuantity": 2,
        "jobNo": "GI022960",
        "startingDate": "2025-08-01T00:00:00",
    },
    {
        "$type": "GilbertAPI.Model.MRP.InProd, GilbertAPI",
        "itemNo": "8315108",
        "description": "DOUILLE DE FIXATION",
        "qtyCanApply": 1,
        "binCode": "P349",
        "name": "M172792",
        "dateGet": "2025-08-07T11:33:00",
    },
]


@pytest.mark.asyncio
async def test_generate_excel_from_sample_data(monkeypatch, tmp_path):
    service = MissingPartsReportService()

    async def fake_fetch(self, sales_order_no: str):
        assert sales_order_no == "GI022960"
        return SAMPLE_OUT_ROWS, SAMPLE_IN_ROWS

    monkeypatch.setattr(MissingPartsReportService, "_fetch_mrp_data", fake_fetch)
    content = await service.generate_excel("gi022960")

    output_path = tmp_path / "missing_parts.xlsx"
    output_path.write_bytes(content)
    workbook = load_workbook(output_path)
    sheet = workbook["Export"]

    parent_row = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value == "M176567":
            parent_row = row
            break

    assert parent_row is not None
    assert sheet.cell(row=parent_row, column=3).value == 8218725
    assert sheet.cell(row=parent_row, column=11).value == 3

    child_items = []
    row_idx = parent_row + 1
    while row_idx <= sheet.max_row:
        if sheet.cell(row=row_idx, column=2).value:
            break
        value = sheet.cell(row=row_idx, column=14).value
        if value not in (None, "", "Item"):
            child_items.append(str(value))
        row_idx += 1

    assert set(child_items) == {"8218725", "622411", "8315108"}

    fill_color = sheet.cell(row=parent_row, column=2).fill.fgColor.value
    assert fill_color in {"FFE1E8D3", "00E1E8D3"}


@pytest.mark.asyncio
async def test_missing_parts_endpoint_streams_excel(monkeypatch):
    sample_bytes = b"excel-bytes"
    mock_generate = AsyncMock(return_value=sample_bytes)
    monkeypatch.setattr("app.api.v1.toolkit.mrp.missing_parts_service.generate_excel", mock_generate)

    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
        response = await client.get("/api/v1/mrp/sales-orders/GI022960/missing-parts/export")

    assert response.status_code == 200
    assert response.headers["content-disposition"].startswith('attachment; filename="missing-parts-GI022960')
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.content == sample_bytes
    mock_generate.assert_awaited_once_with("GI022960")
