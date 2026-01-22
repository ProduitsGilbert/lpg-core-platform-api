"""Build a missing-parts Excel report for a sales order."""

from __future__ import annotations

import asyncio
import io
import json
import logging
import re
from datetime import datetime
from typing import Any, Dict, Iterable, List, Tuple

import httpx
import logfire
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill

from app.settings import settings

logger = logging.getLogger(__name__)


class MissingPartsReportService:
    """Fetch Basic MRP data and render a human-friendly Excel report."""

    def __init__(self) -> None:
        self._base_url = settings.toolkit_base_url.rstrip("/")
        self._headers = {"accept": "*/*"}
        self._timeout = settings.request_timeout

    async def generate_excel(self, sales_order_no: str) -> bytes:
        """
        Fetch MRP data for the sales order and return the Excel bytes.
        """
        normalized = self._normalize_sales_order_no(sales_order_no)
        out_rows, in_rows = await self._fetch_mrp_data(normalized)
        return self._build_excel(normalized, out_rows, in_rows)

    async def _fetch_mrp_data(self, sales_order_no: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        filter_text = self._build_filter(sales_order_no)
        out_task = asyncio.create_task(
            self._fetch_direction(
                "Out",
                params={"filter": filter_text, "type": "-1"},
                span_name="basic_mrp.out",
                sales_order_no=sales_order_no,
            )
        )
        in_task = asyncio.create_task(
            self._fetch_direction(
                "In",
                params={"filter": filter_text, "type": "-1"},
                span_name="basic_mrp.in",
                sales_order_no=sales_order_no,
            )
        )
        return await asyncio.gather(out_task, in_task)

    def _build_filter(self, sales_order_no: str) -> str:
        """Create the upstream filter payload for the given sales order."""
        payload = {"filters": [{"field": "JobNo", "operator": "contains", "value": sales_order_no}]}
        return json.dumps(payload, separators=(",", ":"))

    @staticmethod
    def _normalize_sales_order_no(sales_order_no: str) -> str:
        normalized = (sales_order_no or "").strip().upper()
        if not normalized:
            raise ValueError("sales_order_no is required")
        return normalized

    async def _fetch_direction(
        self,
        direction: str,
        *,
        params: Dict[str, str],
        span_name: str,
        sales_order_no: str,
    ) -> List[Dict[str, Any]]:
        if not self._base_url:
            raise httpx.RequestError("Toolkit base URL is not configured")

        url = f"{self._base_url}/api/mrp/BasicMRP/{direction}"
        span_kwargs = {
            "url": url,
            "direction": direction,
            "sales_order_no": sales_order_no,
        }

        try:
            with logfire.span(span_name, **span_kwargs):
                async with httpx.AsyncClient(timeout=self._timeout) as client:
                    response = await client.get(url, headers=self._headers, params=params or None)
                response.raise_for_status()
                payload = response.json()
        except httpx.HTTPStatusError as exc:
            status_code = exc.response.status_code if exc.response else None
            body = exc.response.text[:500] if exc.response and exc.response.text else ""
            logger.error(
                "Basic MRP %s endpoint returned HTTP error",
                direction,
                extra={"status_code": status_code, "body": body, **span_kwargs},
            )
            raise
        except httpx.RequestError as exc:
            logger.error(
                "Failed to reach Basic MRP %s endpoint",
                direction,
                extra={"error": str(exc), **span_kwargs},
            )
            raise
        except ValueError as exc:
            logger.error("Invalid JSON returned from Basic MRP %s endpoint", direction, exc_info=True)
            raise

        values = payload.get("values") or payload.get("value") or []
        if not isinstance(values, list):
            logger.warning(
                "Unexpected payload format from Basic MRP %s endpoint",
                direction,
                extra={"payload_type": type(values), **span_kwargs},
            )
            return []
        return values

    def _build_excel(
        self,
        sales_order_no: str,
        out_rows: List[Dict[str, Any]],
        in_rows: List[Dict[str, Any]],
    ) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Export"

        self._build_export_sheet(worksheet, sales_order_no, out_rows, in_rows)
        self._autosize_columns(worksheet)

        # Raw data tabs help local validation when the user opens the workbook.
        self._add_raw_sheet(workbook.create_sheet("MRP Out"), out_rows)
        self._add_raw_sheet(workbook.create_sheet("MRP In"), in_rows)

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def _build_export_sheet(
        self,
        worksheet,
        sales_order_no: str,
        out_rows: List[Dict[str, Any]],
        in_rows: List[Dict[str, Any]],
    ) -> None:
        title_fill = PatternFill("solid", fgColor="6B8E23")
        group_fill_complete = PatternFill("solid", fgColor="A6BB7B")
        group_fill_incomplete = PatternFill("solid", fgColor="E1E8D3")
        child_gap_fill = PatternFill("solid", fgColor="B4B4B4")
        header_font = Font(bold=True)

        worksheet.merge_cells("B2:H2")
        worksheet["B2"] = (
            f"{sales_order_no.lower()}    Genere en date de:  {datetime.now():%Y-%m-%d %H:%M:%S}"
        )
        for col in range(2, 9):
            worksheet.cell(row=2, column=col).fill = title_fill

        parent_headers = [
            "Mo",
            "Item",
            "QtyOnMO",
            "QtyRemaining",
            "Type",
            "Rev",
            "Description",
            "Palette",
            "Note",
            "QtyItemBO",
            "Job",
            "DueDate",
        ]
        for offset, header in enumerate(parent_headers):
            cell = worksheet.cell(row=4, column=2 + offset, value=header)
            cell.fill = title_fill
            cell.font = header_font

        child_headers = [
            "Item",
            "Description",
            "ReqQty",
            "WipQty",
            "BinNo",
            "Origin_POMO",
            "WC_Date",
            "Qty",
            "Job",
            "LastMove",
        ]

        in_prod_by_no = self._index_in_prod_rows(in_rows)
        out_groups = self._group_out_rows(out_rows)
        group_keys = self._sorted_group_keys(in_prod_by_no, out_groups)

        row_idx = 5
        for prod_order_no in group_keys:
            child_rows = out_groups.get(prod_order_no, [])
            parent_row = in_prod_by_no.get(prod_order_no)
            parent_values = self._build_parent_values(
                prod_order_no,
                parent_row,
                child_rows,
                sales_order_no,
            )
            child_entries = [self._build_child_entry(row) for row in child_rows]
            group_complete = all(entry["missing_qty"] <= 0 for entry in child_entries) if child_entries else True
            group_fill = group_fill_complete if group_complete else group_fill_incomplete

            for offset, value in enumerate(parent_values):
                cell = worksheet.cell(row=row_idx, column=2 + offset, value=value)
                cell.fill = group_fill

            for offset, header in enumerate(child_headers):
                cell = worksheet.cell(row=row_idx, column=14 + offset, value=header)
                cell.fill = group_fill
                cell.font = header_font

            row_idx += 1

            for entry in child_entries:
                for col in range(2, 14):
                    worksheet.cell(row=row_idx, column=col).fill = child_gap_fill

                for offset, value in enumerate(entry["values"]):
                    cell = worksheet.cell(row=row_idx, column=14 + offset, value=value)
                    cell.fill = group_fill

                row_idx += 1

            row_idx += 2

    def _index_in_prod_rows(self, in_rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        in_prod_by_no: Dict[str, Dict[str, Any]] = {}
        for row in in_rows:
            if not self._is_in_prod(row):
                continue
            key = row.get("prodOrderNo") or row.get("name")
            if key and key not in in_prod_by_no:
                in_prod_by_no[key] = row
        return in_prod_by_no

    def _group_out_rows(self, out_rows: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        groups: Dict[str, List[Dict[str, Any]]] = {}
        for row in out_rows:
            key = row.get("prodOrderNo") or row.get("name")
            if not key:
                continue
            groups.setdefault(key, []).append(row)
        return groups

    def _sorted_group_keys(
        self,
        in_prod_by_no: Dict[str, Dict[str, Any]],
        out_groups: Dict[str, List[Dict[str, Any]]],
    ) -> List[str]:
        keys = {**in_prod_by_no, **out_groups}.keys()

        def sort_key(key: str):
            parent = in_prod_by_no.get(key, {})
            dt = self._parse_dt_sort(
                parent.get("startingDate")
                or parent.get("earliestRequired")
                or parent.get("endingDate")
            )
            return (dt or datetime.max, key)

        return sorted(keys, key=sort_key)

    def _build_parent_values(
        self,
        prod_order_no: str,
        parent_row: Dict[str, Any] | None,
        child_rows: List[Dict[str, Any]],
        sales_order_no: str,
    ) -> List[Any]:
        item_no = parent_row.get("itemNo") if parent_row else ""
        description = ""
        job_no = sales_order_no
        qty_on_mo: Any = ""
        qty_remaining: Any = ""
        due_date: Any = ""
        if parent_row:
            description = parent_row.get("description") or parent_row.get("descriptionItemCard") or ""
            job_no = parent_row.get("jobNo") or sales_order_no
            qty_on_mo = self._coerce_number(parent_row.get("quantity") or parent_row.get("qtyCanApply") or 0)
            qty_remaining = self._coerce_number(
                parent_row.get("remainingQuantity")
                or parent_row.get("qtyDisponible")
                or qty_on_mo
            )
            due_date = self._parse_dt(
                parent_row.get("startingDate")
                or parent_row.get("earliestRequired")
                or parent_row.get("endingDate")
            )

        return [
            prod_order_no,
            self._format_item_no(item_no),
            qty_on_mo,
            qty_remaining,
            "",
            "",
            description,
            "",
            "",
            len(child_rows),
            job_no,
            due_date,
        ]

    def _build_child_entry(self, row: Dict[str, Any]) -> Dict[str, Any]:
        required_qty = self._coerce_number(
            row.get("qtyToFill")
            or row.get("quantityTotal")
            or row.get("qty")
            or row.get("remainingQuantity")
            or 0
        )
        missing_qty = self._missing_quantity(row)
        attribution = row.get("attributionFromIn") or []
        has_supply = self._sum_attribution(attribution) > 0
        is_prod_order_supply = str(row.get("replenishmentSystem") or "").lower().startswith("prod")

        bin_code = row.get("binCode") or ""
        if is_prod_order_supply:
            bin_value = ""
        elif not has_supply and bin_code:
            bin_value = "?"
        else:
            bin_value = bin_code

        origin = self._format_origin(row, attribution, has_supply)
        job = ""
        last_move = ""
        qty = 0
        wc_date = ""
        if is_prod_order_supply:
            first_attr = next((entry for entry in attribution if isinstance(entry, dict)), {})
            job = first_attr.get("job") or ""
            last_move = self._parse_dt(first_attr.get("dateRequiredOrAvailable"))
            qty = self._coerce_number(row.get("lotSize") or 0)

        values = [
            self._format_item_no(row.get("itemNo")),
            row.get("description") or row.get("descriptionItemCard") or "",
            required_qty,
            0,
            bin_value,
            origin,
            wc_date,
            qty,
            job,
            last_move,
        ]

        return {
            "values": [self._sanitize_value(value) for value in values],
            "missing_qty": missing_qty,
        }

    def _format_origin(
        self,
        row: Dict[str, Any],
        attribution: Iterable[Any],
        has_supply: bool,
    ) -> str:
        raw = str(row.get("mrpTakeFrom") or "").strip()
        if raw:
            return self._normalize_origin_text(raw)

        chunks: List[str] = []
        for entry in attribution:
            if not isinstance(entry, dict):
                continue
            qty = self._coerce_number(entry.get("qty"))
            name = entry.get("name") or entry.get("type") or "Source"
            chunks.append(f"{self._format_quantity(qty)}x {name}")
        if chunks:
            return "; ".join(chunks)

        return "VOIR MRP" if not has_supply else ""

    @staticmethod
    def _normalize_origin_text(text: str) -> str:
        text = text.strip().rstrip(";")
        text = re.sub(r";\\s*", "; ", text)
        text = re.sub(r"(\\d+(?:\\.\\d+)?)x(\\S)", r"\\1x \\2", text)
        return " ".join(text.split())

    @staticmethod
    def _format_quantity(value: float) -> str:
        if value.is_integer():
            return str(int(value))
        return f"{value:g}"

    @staticmethod
    def _format_item_no(value: Any) -> Any:
        if value is None:
            return ""
        text = str(value).strip()
        if text.isdigit():
            return int(text)
        return text

    @staticmethod
    def _is_in_prod(row: Dict[str, Any]) -> bool:
        return str(row.get("$type") or "").endswith("InProd, GilbertAPI")

    @staticmethod
    def _parse_dt_sort(value: Any):
        if not value:
            return None
        if isinstance(value, datetime):
            return value
        text = str(value)
        if text.endswith("Z"):
            text = text.replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(text)
            if parsed.tzinfo is not None:
                return parsed.replace(tzinfo=None)
            return parsed
        except ValueError:
            return None

    def _build_missing_rows(
        self,
        out_rows: List[Dict[str, Any]],
        in_rows: List[Dict[str, Any]],
        sales_order_no: str,
    ) -> Iterable[List[Any]]:
        for row in out_rows:
            missing_qty = self._missing_quantity(row)

            qty_required = self._coerce_number(
                row.get("qtyToFill")
                or row.get("quantityTotal")
                or row.get("qty")
                or row.get("remainingQuantity")
                or 0
            )
            supplied_qty = self._sum_attribution(row.get("attributionFromIn"))
            inbound_match_text, earliest_inbound = self._describe_inbound(row.get("itemNo"), in_rows)

            cells = [
                sales_order_no,
                row.get("name") or row.get("prodOrderNo") or row.get("noHeader") or "",
                row.get("itemNo") or "",
                row.get("description") or row.get("descriptionItemCard") or "",
                self._parse_dt(row.get("needDate") or row.get("orderDate")),
                qty_required,
                supplied_qty,
                missing_qty,
                row.get("binCode") or "",
                row.get("mrpTakeFrom") or "",
                inbound_match_text,
                earliest_inbound,
            ]
            yield [self._sanitize_value(value) for value in cells]

    def _sum_attribution(self, attribution: Any) -> float:
        total = 0.0
        if not attribution:
            return total

        for entry in attribution:
            if isinstance(entry, dict):
                total += self._coerce_number(entry.get("qty"))
        return total

    def _describe_inbound(
        self,
        item_no: Any,
        in_rows: List[Dict[str, Any]],
    ) -> Tuple[str, Any]:
        matches = [row for row in in_rows if str(row.get("itemNo")) == str(item_no)]
        if not matches:
            return "", ""

        chunks: List[str] = []
        earliest_date = None
        for row in matches:
            qty_available = (
                row.get("qtyDisponible")
                or row.get("qtyCanApply")
                or row.get("quantity")
                or row.get("remainingQuantity")
                or 0
            )
            source = row.get("name") or row.get("$type") or "Inbound"
            chunk = f"{source} ({self._coerce_number(qty_available)})"
            if row.get("binCode"):
                chunk += f" @ {row['binCode']}"
            chunks.append(chunk)

            dt_value = row.get("expectedReceiptDate") or row.get("earliestRequired") or row.get("dateGet")
            parsed_dt = self._parse_dt(dt_value)
            if parsed_dt and (earliest_date is None or parsed_dt < earliest_date):
                earliest_date = parsed_dt

        return "; ".join(chunks), earliest_date

    def _missing_quantity(self, row: Dict[str, Any]) -> float:
        qty_to_fill = self._coerce_number(row.get("qtyToFill") or row.get("quantityTotal") or row.get("qty") or 0)
        qty_filled = self._coerce_number(row.get("qtyFilled") or 0)
        base_missing = max(qty_to_fill - qty_filled, 0.0)

        remainder_fields = [
            row.get("qtyRemainingFIll"),
            row.get("qtyRemainingFill"),
            row.get("remainingQuantity"),
        ]
        remainder = max(self._coerce_number(value) for value in remainder_fields)
        if base_missing <= 0 and remainder > 0:
            base_missing = remainder

        supplied_qty = self._sum_attribution(row.get("attributionFromIn"))
        if supplied_qty:
            return max(base_missing - supplied_qty, 0.0)
        return max(base_missing, remainder)

    def _parse_dt(self, value: Any):
        """Return datetime for Excel or an empty string when missing."""
        if not value:
            return ""

        text = str(value)
        if text.endswith("Z"):
            text = text.replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(text)
            if parsed.tzinfo is not None:
                return parsed.replace(tzinfo=None)
            return parsed
        except ValueError:
            return text

    def _coerce_number(self, value: Any) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    def _style_header(self, worksheet) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.freeze_panes = "A2"

    def _autosize_columns(self, worksheet) -> None:
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 60)

    def _add_raw_sheet(self, worksheet, rows: List[Dict[str, Any]]) -> None:
        if not rows:
            worksheet.append(["No data"])
            self._style_header(worksheet)
            self._autosize_columns(worksheet)
            return

        # Stable header ordering for readability in the spreadsheet.
        keys: List[str] = sorted({key for row in rows for key in row.keys()})
        worksheet.append(keys)
        for row in rows:
            worksheet.append([self._cell_value(row.get(key, "")) for key in keys])
        self._style_header(worksheet)
        self._autosize_columns(worksheet)

    def _cell_value(self, value: Any) -> Any:
        return self._sanitize_value(value)

    def _sanitize_value(self, value: Any) -> Any:
        if isinstance(value, (list, dict)):
            try:
                value = json.dumps(value)
            except TypeError:
                value = str(value)

        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value
